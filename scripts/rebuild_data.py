#!/usr/bin/env python3
"""Regenerate the M / PM / P / MG data blocks in stats.html from the Google
Sheets export (Log tuần = event log, Info = canonical roster).

Usage:
    python3 scripts/rebuild_data.py --xlsx <path> [--html stats.html] [--dry-run]
"""
import argparse
import json
import re
import warnings
from collections import defaultdict
from pathlib import Path

import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parent
DEFAULT_XLSX = (
    "/tmp/claude-0/-home-user-S7CN/e0b2620f-aa50-5bbf-b72e-abe3d3c5249f/"
    "scratchpad/sheet.xlsx"
)

# Sheet name -> "Dương Chí Hoàng" is displayed elsewhere as "D.Chí Hoàng".
# "Khanh" là lỗi gõ thiếu dấu của "Khánh" (bạn Duyên) ở Tuần 30 — cùng một
# người, gộp lại để không tách thành hai cầu thủ.
DISPLAY = {"Dương Chí Hoàng": "D.Chí Hoàng", "Khanh": "Khánh"}

RESULT_MAP = {"Thắng": "W", "Hòa": "D", "Thua": "L"}
GOAL_TYPE_MAP = {"Thường": "normal", "Pen (Vào)": "pen", "Sút phạt": "fk"}
HALF_MAP = {"H1": 1, "H2": 2, None: None}


def disp(name):
    """Apply the display-name override, pass through everything else."""
    if name is None:
        return name
    return DISPLAY.get(name, name)


def fmt_no(value):
    """Số áo -> string without the trailing '.0' ('' if blank)."""
    if value is None or value == "":
        return ""
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return str(value)
    return str(value)


def load_roster(wb):
    """Read the canonical roster from the `Info` sheet.

    Row 0 is the header. col0 TT / col1 Họ và tên / col2 Biệt danh /
    col3 Số áo / col4 Vị trí (all 0-based).

    The sheet has a run of rows (Info rows 20-23 in the current export)
    that carry a leftover Số áo value but no Họ và tên / Biệt danh - these
    are not player rows and are skipped. Reading stops at the first row
    where Họ và tên, Biệt danh, Số áo and Vị trí are ALL empty (the real
    end-of-roster blank block), which is what actually yields the
    documented 18 main + 15 sub = 33 rows.
    """
    ws = wb["Info"]
    main, subs = [], []
    for r in range(2, ws.max_row + 1):
        tt = ws.cell(row=r, column=1).value
        hoten = ws.cell(row=r, column=2).value
        biet = ws.cell(row=r, column=3).value
        no = ws.cell(row=r, column=4).value
        pos = ws.cell(row=r, column=5).value

        if hoten is None and biet is None and no is None and pos is None:
            break  # fully blank row => end of roster block
        if biet is None:
            continue  # orphan row, no join key, not a real player

        entry = {"n": disp(biet), "p": pos or "", "no": fmt_no(no)}
        if isinstance(tt, (int, float)):
            entry["_tt"] = int(tt)
            entry["t"] = "main"
            main.append(entry)
        elif tt == "Sub":
            entry["t"] = "sub"
            subs.append(entry)
        # else: unrecognised TT value, skip

    main.sort(key=lambda e: e["_tt"])
    for e in main:
        del e["_tt"]
    return main + subs  # canonical roster order: main by TT asc, then subs


def load_weeks(wb):
    """Group `Log tuần` rows by week number. Row order in the sheet is NOT
    reliable (a week's rows can be split across the file), so we group
    strictly by the `Tuần N` value in col0, not by row position.
    """
    ws = wb["Log tuần"]
    weeks = defaultdict(list)
    for r in range(2, ws.max_row + 1):
        wk = ws.cell(row=r, column=1).value
        if wk is None:
            continue
        row = [ws.cell(row=r, column=c).value for c in range(1, 15)]
        weeks[wk].append(row)
    return weeks


def wk_num(wk_label):
    return int(wk_label.split()[1])


def load_live_links(wb):
    """Đọc link xem lại trận từ cột O sheet 'Thống kê match'.

    Trả về {wk_num: url}. Chỉ nhận hàng có kết quả Thắng/Hòa/Thua — cột O
    từng có link nằm nhầm ở hàng tuần Nghỉ (Tuần 5), tuần đó không có trận
    nên không thể gắn link vào đâu.
    """
    if "Thống kê match" not in wb.sheetnames:
        return {}
    ws = wb["Thống kê match"]
    links = {}
    for r in range(1, ws.max_row + 1):
        wk_cell = ws.cell(row=r, column=1).value
        if not wk_cell or not str(wk_cell).strip().startswith("Tuần"):
            continue
        result = ws.cell(row=r, column=10).value  # J = Kết quả
        url = ws.cell(row=r, column=15).value     # O = link xem lại
        if not url or str(result).strip() not in RESULT_MAP:
            continue
        url = str(url).strip()
        if not url:
            continue
        if not url.lower().startswith(("http://", "https://")):
            url = "https://" + url
        try:
            links[wk_num(str(wk_cell))] = url
        except Exception:
            continue
    return links


def build(xlsx_path):
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")  # harmless pivotCache UserWarning
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    roster = load_roster(wb)
    weeks = load_weeks(wb)
    live_links = load_live_links(wb)

    played = []  # (wk_num, rows) for weeks that were actually played
    for wk_label, rows in weeks.items():
        if all(r[9] == "Nghỉ" for r in rows):
            continue  # week off, skip entirely
        played.append((wk_num(wk_label), rows))
    played.sort(key=lambda t: t[0])

    # ---- M ----
    M = []
    for wn, rows in played:
        head = rows[0]
        dt_val = head[1]
        dt = f"{dt_val.day:02d}/{dt_val.month:02d}"
        entry = {
            "wk": wn,
            "dt": dt,
            "vs": head[3],
            "gf": int(head[6]),
            "ga": int(head[7]),
            "r": RESULT_MAP[head[8]],
            "v": head[4],
        }
        if wn in live_links:
            entry["lv"] = live_links[wn]
        M.append(entry)
    week_order = [m["wk"] for m in M]

    # ---- attendance + goals-per-player-per-week (for PM) + MG ----
    # "Implied attendance": the sheet has weeks where a player has a
    # Bàn thắng / assist / Bàn thua(gk) record but NO Điểm danh row. You
    # cannot score, assist or keep goal without playing, so such a
    # player is treated as present that week even without an explicit
    # attendance row (e.g. T20 "Lâm" scored with no Điểm danh row; T19
    # "Thủ môn mới" conceded 3 as GK with no Điểm danh row).
    attendance = defaultdict(set)  # wk -> set(player names)
    implied_attendance = defaultdict(list)  # wk -> [(name, role)] implied-only cases
    goals_by_player = defaultdict(lambda: defaultdict(int))  # wk -> name -> count
    MG = {}
    for wn, rows in played:
        s_list, c_list, ps_list = [], [], []
        for row in rows:
            kind = row[9]
            if kind == "Điểm danh":
                attendance[wn].add(disp(row[10]))
            elif kind == "Bàn thắng":
                name = disp(row[10])
                goals_by_player[wn][name] += 1
                entry = {"sc": name}
                if row[11]:
                    entry["as"] = disp(row[11])
                entry["t"] = GOAL_TYPE_MAP.get(row[13], "normal")
                entry["h"] = HALF_MAP.get(row[12], None)
                s_list.append(entry)
            elif kind == "Bàn thua":
                entry = {"gk": disp(row[11]), "h": HALF_MAP.get(row[12], None)}
                if row[13] in ("Pen", "Pen (Vào)"):
                    entry["pen"] = True
                c_list.append(entry)
            elif kind == "Cản Pen":
                ps_list.append({"n": disp(row[11]), "h": HALF_MAP.get(row[12], None)})
        MG[wn] = {"s": s_list, "c": c_list, "ps": ps_list}

    # Apply the implied-attendance union: scorer / assist / conceding GK
    # each imply presence, even without a Điểm danh row. Track which ones
    # were NOT already covered by an explicit attendance row, for the
    # warning report in check_stats.py.
    for wn, mg in MG.items():
        for g in mg["s"]:
            for role, key in (("ghi bàn", "sc"), ("kiến tạo", "as")):
                name = g.get(key)
                if not name:
                    continue
                if name not in attendance[wn]:
                    implied_attendance[wn].append((name, role))
                attendance[wn].add(name)
        for c in mg["c"]:
            name = c["gk"]
            if name not in attendance[wn]:
                implied_attendance[wn].append((name, "thủ môn"))
            attendance[wn].add(name)
        for ev in mg["ps"]:
            name = ev["n"]
            if name not in attendance[wn]:
                implied_attendance[wn].append((name, "cản pen"))
            attendance[wn].add(name)

    # ---- P : roster players with >=1 match, in canonical roster order ----
    P = []
    for entry in roster:
        name = entry["n"]
        if any(name in attendance[wn] for wn in week_order):
            P.append({"n": name, "p": entry["p"], "no": entry["no"], "t": entry["t"]})

    # Khách chỉ có trong 'Log tuần' mà chưa có dòng nào trong 'Info' (ví dụ
    # 'Bạn mới' ghi bàn ở T14). Bỏ họ đi thì bàn thắng nằm trong MG nhưng
    # không quy được cho ai trong P/PM -> lệch tổng bàn. Đưa vào cuối roster
    # dạng sub, không số áo / không vị trí, và báo ra để anh bổ sung vào Info.
    known = {p["n"] for p in P} | {e["n"] for e in roster}
    log_only = []
    for wn in week_order:
        for name in sorted(attendance[wn]):
            if name not in known:
                known.add(name)
                log_only.append(name)
                P.append({"n": name, "p": "", "no": "", "t": "sub"})

    # ---- PM ----
    PM = {}
    for p in P:
        name = p["n"]
        arr = []
        for wn in week_order:
            if name not in attendance[wn]:
                arr.append(None)
            else:
                arr.append(goals_by_player[wn].get(name, 0))
        PM[name] = arr

    return M, PM, P, MG, implied_attendance, log_only


# ---------------------------------------------------------------------------
# JS rendering
# ---------------------------------------------------------------------------

def js_str(s):
    return json.dumps(s, ensure_ascii=False)


def render_M(M):
    lines = ["const M=["]
    for m in M:
        lv = ",lv:" + js_str(m["lv"]) if m.get("lv") else ""
        lines.append(
            "  {{wk:{wk},dt:{dt},vs:{vs},gf:{gf},ga:{ga},r:{r},v:{v}{lv}}},".format(
                wk=m["wk"],
                dt=js_str(m["dt"]),
                vs=js_str(m["vs"]),
                gf=m["gf"],
                ga=m["ga"],
                r=js_str(m["r"]),
                v=js_str(m["v"]),
                lv=lv,
            )
        )
    lines.append("];")
    return "\n".join(lines)


def render_MG(MG):
    lines = ["const MG={"]
    for wn in sorted(MG.keys()):
        mg = MG[wn]
        s_parts = []
        for g in mg["s"]:
            fields = [f"sc:{js_str(g['sc'])}"]
            if "as" in g:
                fields.append(f"as:{js_str(g['as'])}")
            fields.append(f"t:{js_str(g['t'])}")
            fields.append(f"h:{'null' if g['h'] is None else g['h']}")
            s_parts.append("{" + ",".join(fields) + "}")
        c_parts = []
        for c in mg["c"]:
            h = "null" if c["h"] is None else c["h"]
            pen = ",pen:true" if c.get("pen") else ""
            c_parts.append(f"{{gk:{js_str(c['gk'])},h:{h}{pen}}}")
        ps_parts = []
        for ev in mg["ps"]:
            h = "null" if ev["h"] is None else ev["h"]
            ps_parts.append(f"{{n:{js_str(ev['n'])},h:{h}}}")
        lines.append(
            "  {wn}:{{s:[{s}],c:[{c}],ps:[{ps}]}},".format(
                wn=wn, s=",".join(s_parts), c=",".join(c_parts), ps=",".join(ps_parts)
            )
        )
    lines.append("};")
    return "\n".join(lines)


def render_P(P):
    lines = ["const P=["]
    for p in P:
        lines.append(
            "  {{n:{n},p:{p},no:{no},t:{t}}},".format(
                n=js_str(p["n"]), p=js_str(p["p"]), no=js_str(p["no"]), t=js_str(p["t"])
            )
        )
    lines.append("];")
    return "\n".join(lines)


def render_PM(PM):
    lines = ["const PM={"]
    items = list(PM.items())
    for i, (name, arr) in enumerate(items):
        arr_str = ",".join("null" if v is None else str(v) for v in arr)
        comma = "," if i < len(items) - 1 else ""
        lines.append(f"  {js_str(name)}:[{arr_str}]{comma}")
    lines.append("};")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# stats.html rewriting
# ---------------------------------------------------------------------------

def find_block(text, const_name):
    """Find `const <name>=...;` where the value is a [...] or {...} literal,
    scanning by bracket/brace depth from the opening char. Returns
    (start_index, end_index_exclusive) spanning the whole statement
    including the trailing ';'.
    """
    pattern = re.compile(r"^const " + re.escape(const_name) + r"=", re.MULTILINE)
    m = pattern.search(text)
    if not m:
        raise ValueError(f"const {const_name}= not found in html")
    start = m.start()
    open_pos = m.end()
    open_ch = text[open_pos]
    if open_ch == "[":
        close_ch = "]"
    elif open_ch == "{":
        close_ch = "}"
    else:
        raise ValueError(f"Unexpected literal opening for {const_name}: {open_ch!r}")

    depth = 0
    i = open_pos
    in_str = None
    while i < len(text):
        ch = text[i]
        if in_str:
            if ch == "\\":
                i += 2
                continue
            if ch == in_str:
                in_str = None
            i += 1
            continue
        if ch in ("'", '"'):
            in_str = ch
            i += 1
            continue
        if ch == open_ch:
            depth += 1
        elif ch == close_ch:
            depth -= 1
            if depth == 0:
                # expect a ';' right after
                j = i + 1
                if j < len(text) and text[j] == ";":
                    return start, j + 1
                return start, j
        i += 1
    raise ValueError(f"Unterminated literal for {const_name}")


def replace_block(text, const_name, new_block):
    start, end = find_block(text, const_name)
    return text[:start] + new_block + text[end:]


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=DEFAULT_XLSX)
    ap.add_argument("--html", default=str(REPO_ROOT / "stats.html"))
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--force", action="store_true",
                    help="ghi đè kể cả khi phát hiện link xem lại sẽ bị mất")
    args = ap.parse_args()

    M, PM, P, MG, implied_attendance, log_only = build(args.xlsx)

    block_M = render_M(M)
    block_MG = render_MG(MG)
    block_P = render_P(P)
    block_PM = render_PM(PM)

    # ---- summary ----
    n_main = sum(1 for p in P if p["t"] == "main")
    n_sub = sum(1 for p in P if p["t"] == "sub")
    total_pm_goals = sum(v for arr in PM.values() for v in arr if v)
    total_mg_goals = sum(len(mg["s"]) for mg in MG.values())

    mismatches = []
    m_by_wk = {m["wk"]: m for m in M}
    for wn, mg in MG.items():
        m = m_by_wk.get(wn)
        if m is None:
            continue
        if len(mg["s"]) != m["gf"]:
            mismatches.append(f"  wk{wn}: len(MG.s)={len(mg['s'])} != M.gf={m['gf']}")
        if len(mg["c"]) != m["ga"]:
            mismatches.append(f"  wk{wn}: len(MG.c)={len(mg['c'])} != M.ga={m['ga']}")

    print(f"Matches: {len(M)}")
    print(f"Players: {len(P)} (main={n_main}, sub={n_sub})")
    print(f"Total goals in PM: {total_pm_goals}")
    print(f"Total goals in MG: {total_mg_goals}")
    total_gf = sum(m["gf"] for m in M)
    total_ga = sum(m["ga"] for m in M)
    print(f"Total M.gf: {total_gf}  Total M.ga: {total_ga}")
    if mismatches:
        print("Per-week len(MG.s)/len(MG.c) vs M.gf/M.ga mismatches:")
        print("\n".join(mismatches))
    else:
        print("No per-week MG/M count mismatches.")

    if log_only:
        print("Cầu thủ chỉ có trong 'Log tuần', CHƯA có dòng trong sheet 'Info'"
              " (đã tự thêm vào cuối roster dạng sub, nên bổ sung vào Info):")
        for name in log_only:
            print(f"  {name}")

    if implied_attendance:
        print("Implied-attendance warnings (no Điểm danh row, inferred from goal/assist/GK record):")
        for wn in sorted(implied_attendance):
            for name, role in implied_attendance[wn]:
                print(f"  T{wn} {name} ({role}) thiếu dòng điểm danh")

    if args.dry_run:
        print("\n--- const M ---")
        print(block_M)
        print("\n--- const MG ---")
        print(block_MG)
        print("\n--- const P ---")
        print(block_P)
        print("\n--- const PM ---")
        print(block_PM)
        return

    html_path = Path(args.html)
    text = html_path.read_text(encoding="utf-8")
    # Chốt chặn: không được âm thầm làm mất link xem lại đang có trong html.
    # Xảy ra khi file xlsx cũ hơn sheet thật (link đã được dời sang tuần khác).
    old_links = dict(re.findall(r'\{wk:(\d+),[^}]*lv:"([^"]+)"', text))
    new_links = {str(m["wk"]): m["lv"] for m in M if m.get("lv")}
    lost = {w: u for w, u in old_links.items() if new_links.get(w) != u}
    if lost:
        print("\n[CẢNH BÁO] link xem lại trong stats.html sẽ bị thay đổi/mất:")
        for w, u in sorted(lost.items(), key=lambda x: int(x[0])):
            print(f"  T{w}: {u}  ->  {new_links.get(w, 'KHÔNG CÒN')}")
        print(f"  html đang có {len(old_links)} link, sheet cho ra {len(new_links)} link.")
        print("  Nếu không cố ý: file xlsx có thể cũ hơn sheet thật. Tải lại sheet rồi chạy lại.")
        if not args.force:
            print("  Đã DỪNG, chưa ghi gì. Thêm --force nếu chắc chắn muốn ghi đè.")
            return

    text = replace_block(text, "M", block_M)
    text = replace_block(text, "MG", block_MG)
    text = replace_block(text, "P", block_P)
    text = replace_block(text, "PM", block_PM)
    html_path.write_text(text, encoding="utf-8")
    print(f"\nWrote updated blocks to {html_path}")


if __name__ == "__main__":
    main()
